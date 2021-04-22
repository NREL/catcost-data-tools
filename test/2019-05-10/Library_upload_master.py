# -*- coding: utf-8 -*-
"""
Created on Thu Jun 14 07:50:03 2018

@author: mjankous
"""

import pandas as pd
import os
import numpy as np
import math
import uuid
import json
import time
from openpyxl import load_workbook
from xlrd import open_workbook

def get_materials_lib(excel_path):
    """
    gets the existing materials library as a pandas dataframe. Will be used to create json file
    """
    ccm_mat_lib_df = pd.read_excel(excel_path,
                                   sheet_name="Materials Library",skiprows=14)
    return ccm_mat_lib_df



def get_icis():
    """
    gathers the ICIS price data in a pandas dataframe to integrate it into the materials library
    """
    icis_df = pd.read_excel("C:\\Users\\mjankous\\Documents\\CCM\\ICIS_price_spreadsheet.xlsx",
                            sheet_name="ICIS_prices", skiprows=10, usecols=[15,6,8,11,12,13,14,18,19])
    icis_df = icis_df.rename(columns={"Full Name":"Material Name",
                                      "Quantity":"Bulk Quote Quantity",
                                      "Corrected Units":"Bulk Quote Units",
                                      "Price Avg":"Bulk Quote Price ($)",
                                      "Year":"Quote Year",
                                      "Unit Price Average":"Unit Price",
                                      "Unit Price Dimensions":"Unit price dimensions"})
    nrows = len(icis_df)
    icis_df['Material Type'] = [np.nan] * nrows
    icis_df['MW (g/mol)'] = [np.nan] * nrows
    icis_df['Concentration (%)'] = [np.nan] * nrows
    icis_df['Lab-Scale Log Fit?'] = [np.nan] * nrows
    icis_df['Lab Quantity 1'] = [np.nan] * nrows
    icis_df['Lab Quantity 2'] = [np.nan] * nrows
    icis_df['Lab Quantity 3'] = [np.nan] * nrows
    icis_df['Lab Quantity 4'] = [np.nan] * nrows
    icis_df['Lab Units'] = [np.nan] * nrows
    icis_df['Lab Price 1'] = [np.nan] * nrows
    icis_df['Lab Price 2'] = [np.nan] * nrows
    icis_df['Lab Price 3'] = [np.nan] * nrows
    icis_df['Lab Price 4'] = [np.nan] * nrows
    icis_df['Quote Source'] = ['ICIS'] * nrows
    icis_df['Quote Access Date'] = ['?'] * nrows
    icis_df['Notes'] = [np.nan] * nrows
    icis_df['Lab Quote Count'] = [np.nan] * nrows #I don't know what this is supposed to represent
    icis_df['Lab Q Model Units 1'] = [np.nan] * nrows
    icis_df['Lab Q Model Units 2'] = [np.nan] * nrows
    icis_df['Lab Q Model Units 3'] = [np.nan] * nrows
    icis_df['Lab Q Model Units 4'] = [np.nan] * nrows
    icis_df['Lab Unit P 1'] = [np.nan] * nrows
    icis_df['Lab Unit P 2'] = [np.nan] * nrows
    icis_df['Lab Unit P 3'] = [np.nan] * nrows
    icis_df['Lab Unit P 4'] = [np.nan] * nrows
    icis_df['Lab Log-Log Slope'] = [np.nan] * nrows
    icis_df['Lab Log-Log Intercept'] = [np.nan] * nrows
    icis_df['Basis Cell'] = [np.nan] * nrows
    icis_df['Lab Forcast Unit Price (Output Mass Units)'] = [np.nan] * nrows
    icis_df['Bulk quote quantity in model units'] = [np.nan] * nrows
    icis_df['Unit price in quote year'] = [np.nan] * nrows
    icis_df['Unit Price'] = [np.nan] * nrows    #comment this out to use original calculations
    icis_df = icis_df[["Material Name",'Material Type','MW (g/mol)','Density (g/mL)',
             'Concentration (%)','Lab-Scale Log Fit?','Lab Quantity 1',
             'Lab Quantity 2','Lab Quantity 3','Lab Quantity 4','Lab Units',
             'Lab Price 1','Lab Price 2','Lab Price 3','Lab Price 4',
             "Bulk Quote Price ($)","Bulk Quote Quantity","Bulk Quote Units",
             "Quote Year",'Quote Source','Quote Access Date','Notes',
             'Lab Quote Count','Lab Q Model Units 1','Lab Q Model Units 2',
             'Lab Q Model Units 3','Lab Q Model Units 4','Lab Unit P 1',
             'Lab Unit P 2','Lab Unit P 3','Lab Unit P 4','Lab Log-Log Slope',
             'Lab Log-Log Intercept','Basis Cell',
             'Lab Forcast Unit Price (Output Mass Units)',
             'Bulk quote quantity in model units','Unit price in quote year',
             "Unit Price","Unit price dimensions"]]
    icis_df['Bulk Quote Units'] = icis_df.apply(lambda row: rename_tons(row), axis=1)
    return icis_df

def rename_tons(row):
    #renames instances of singular ton from ICIS database to match format of current materials library
    unit = row['Bulk Quote Units']
    if unit == 'ton':
        new_unit = 'tons'
    elif unit == 'tonne':
        new_unit = 'tonnes'
    else:
        new_unit = unit
    return new_unit

def gen_id():
    #randomly creates a uuid to be used on a material
    an_id = str(uuid.uuid4())
    return an_id

def date_to_str(row):
    #given the presence of several different date formats, this one normalizes them all to strings
    #this script has been dramatically simplified now that the date format is normalized
    #it simply accounts for the empty rows
    ts = row['Quote Access Date']
#    try:
    dt = ts.to_pydatetime()

#    except AttributeError:
#        dt = ts
    try:
        date_string = dt.strftime('%m/%d/%Y')
    except ValueError: #try-except control sequence used to resolve NaTType error for blank entries in the quote access date column
        date_string = "01/01/" + str(row['Quote Year'])


    ep = int(time.mktime(time.strptime(date_string, '%m/%d/%Y')))  
#    print(date_string)
#    print(ep)     

#    if date_string == "?":
#        date_string_tmp = "01/01/" + str(row['Quote Year'])
#        ep = int(time.mktime(time.strptime(date_string_tmp,'%m/%d/%Y')))
#    elif date_string == "RP153":
#        date_string_tmp = "01/01/" + str(row['Quote Year'])
#        ep = int(time.mktime(time.strptime(date_string_tmp,'%m/%d/%Y')))
#    elif date_string == "01/01/1900":
#        date_string = "01/01/" + str(row['Quote Year'])
#        ep = int(time.mktime(time.strptime(date_string,'%m/%d/%Y')))
#    elif type(date_string) == float:
#        if math.isnan(date_string):
#            date_string = "01/01/" + str(row['Quote Year'])
#            ep = int(time.mktime(time.strptime(date_string,'%m/%d/%Y')))
#    else
#        #print(date_string)
#        ep = int(time.mktime(time.strptime(date_string, '%m/%d/%Y')))
    return ep, date_string

def make_price_dict(entry):
    """
    reformats the data in the dictionaries for each material as a price dictionary following the convention provided
    """
    price_dict = {}
    price_type = "lab-scale-log-fit" if entry["Lab-Scale Log Fit?"] == 'Lab' else "quote"
    #tests if the data in the materials library for a material is lab or bulk quote data
    price_dict["type"] = price_type
    price_dict["year"] = str(entry["Quote Year"])
    
    if price_type == "quote":
        #fills out quote block for quote prices\
        price_dict['bulk_quote_units'] = entry['bulk_quote_units']
        price_dict["quote"] = {}
        price_dict["quote"]["source"] = entry["Quote Source"]
        price_dict["quote"]["price"] = entry["Bulk Quote Price ($)"]
        price_dict["quote"]["quantity"] = entry["Bulk Quote Quantity"]
        price_dict["quote"]["date"] = date_to_str(entry)[1]
        if price_dict['bulk_quote_units'] == 'ton':
            price_dict['bulk_quote_units'] = 'tons'
        elif price_dict['bulk_quote_units'] == 'tonne':
            price_dict['bulk_quote_units'] = 'tonnes'
    #labels material units
   
    else:
        #fills out lab block for lab prices and gives default values for quote
        price_dict['lab_scale_units'] = entry['lab_scale_units']
        price_dict["lab-scale-log-fit"] = {}
        price_dict["lab-scale-log-fit"]["values"] = []
        lab_quote_count = entry["Lab Quote Count"]
        for lq in range(1,lab_quote_count+1):
            #steps through each listed lab price and quantity and adds them to the lab-scale-log-fit values
            lq_dict = {}
            lq_dict["price"] = entry["Lab Price %s" %lq]
            lq_dict["quantity"] = entry["Lab Quantity %s" %lq]
            price_dict["lab-scale-log-fit"]["values"].append(lq_dict)
        if price_dict['lab_scale_units'] == 'ton':
            price_dict['lab_scale_units'] = 'tons'
        elif price_dict['lab_scale_units'] == 'tonne':
            price_dict['lab_scale_units'] = 'tonnes'
            #labels material units
    
    
    
    return price_dict
    



def materials_to_json(excel_path, json_path, complete=False, version="1.0.0", remove_proprietary=True):
    if complete:
        mat_lib_df = get_materials_lib(excel_path)
    else:
        mat_lib_df = pd.read_excel(excel_path, 
                                   sheet_name="Materials Library",skiprows=14,
                                   skip_footer=0)
    nrows = len(mat_lib_df)
    mat_lib_df["version"] = [version] * nrows
    #populates version as 1.0.0 for all materials. may become deprecated
    mat_lib_df = mat_lib_df.rename(columns={"Material Name":"name","Material Type":"type",
                                            "MW (g/mol)":"molecularWeight",
                                            "Density (g/mL)":"density",
                                            "Concentration (%)":"concentration",
                                            "Lab Units":"lab_scale_units",
                                            "Bulk Quote Units":"bulk_quote_units"})
    #changes names of columns to match json
    mat_lib_df = mat_lib_df.drop(["Notes","Basis Cell"], 1)
    mat_lib_df = mat_lib_df[mat_lib_df['Quote Source'] != 'IHS PEP']
    mat_lib_df = mat_lib_df[mat_lib_df['Quote Source'] != 'IHS CEH']
    mat_lib_df = mat_lib_df[mat_lib_df['Quote Source'] != 'IHS PEP quote']
    #mat_lib_df = mat_lib_df[mat_lib_df['Bulk Quote Units'] != 'cyl']
        
    clean_up_lst = ["Bulk Quote Price ($)","Bulk Quote Quantity",'Bulk Quote Units',
                    'Bulk quote quantity in model units','Lab Forecast Unit Price (Output Mass Units)',
                    'Lab Log-Log Intercept','Lab Log-Log Slope','Lab Price 1', 
                    'Lab Price 2','Lab Price 3','Lab Price 4','Lab Q Model Units 1',
                    'Lab Q Model Units 2','Lab Q Model Units 3','Lab Q Model Units 4',
                    'Lab Quantity 1','Lab Quantity 2','Lab Quantity 3', 'Lab Quantity 4',
                    'Lab Quote Count','Lab Unit P 1','Lab Unit P 2','Lab Unit P 3',
                    'Lab Unit P 4','Lab Units','Lab-Scale Log Fit?','Quote Access Date',
                    'Quote Source','Quote Year','Unit price','Unit price dimensions',
                    'Unit price in quote year', 'lab_scale_units', 'bulk_quote_units']
    #list of unnecessary keys removed from json file
    mat_lib_dict = mat_lib_df.to_dict('records')
    #convert df to dict
    mat_id_dict = get_ids('mat_id_dict')
    for entry in mat_lib_dict:
        entry["price"] = make_price_dict(entry)
        #generates price dictionary for each dictionary
        #entry["id"] = gen_id()
        try:
            entry["id"] = mat_id_dict[entry['name']]
        except KeyError:
            entry['id'] = add_id('mat_id_dict',entry['name'])
        #create id for each material
        entry["updatedOn"] = date_to_str(entry)[0]
        for item in clean_up_lst:
            entry.pop(item,None)
            #removes any columns listed in the clean up list
        if type(entry["type"]) == float:
            entry.pop('type', None)
        for key in entry:
            if entry[key] == "":
                entry[key] = None
            if type(entry[key]) == float:
                if math.isnan(entry[key]): 
                    entry[key] = None
                    #change back to none for final version
                    #changes nan values in outer dictionary of each entry into null values
        if 'quote' in entry['price'].keys():
            if type(entry['price']['quote']['date']) == float:
                if math.isnan(entry['price']['quote']['date']):
                    entry['price']['quote']['date'] = None
                #change back to none for final version
                #changes nan values for "date" within the price quote dictionary into null values
    mat_lib_json = json.dumps(mat_lib_dict,indent=2)
    #converts dictionary to json
    if json_path != None:
        with open(json_path,'w') as mat_json_out:
            mat_json_out.write(mat_lib_json)
    return mat_lib_json

def materials_json_to_excel(excel_path, json_path):
    with open(json_path,'r') as f:
        mat_json = f.read()
        mat_lst = json.loads(mat_json)
    d = {'Material Name':[],'Material Type':[], 'MW (g/mol)':[], 'Density (g/mL)':[], 
         'Concentration (%)':[], 'Lab-Scale Log Fit?':[],'Lab Quantity 1':[], 
         'Lab Quantity 2':[], 'Lab Quantity 3':[], 'Lab Quantity 4':[],
         'Lab Price 1':[], 'Lab Price 2':[], 'Lab Price 3':[], 'Lab Price 4':[],
         'Lab Units':[], 'Bulk Quote Price ($)':[],'Bulk Quote Quantity':[],
         'Bulk Quote Units':[], 'Quote Source':[], 'Quote Access Date':[], 'Quote Year':[]}
    for entry in mat_lst:
        d['Material Name'].append(entry['name'])
        if 'type' in entry.keys():
            d['Material Type'].append(entry['type'])
        else:
            d['Material Type'].append(None)
        d['Density (g/mL)'].append(entry['density'])
        d['MW (g/mol)'].append(entry['molecularWeight'])
        d['Concentration (%)'].append(entry['concentration'])
        if 'bulk_quote_units' in entry['price'].keys():
            if entry['price']['bulk_quote_units'] == 'tons':
                d['Bulk Quote Units'].append('ton')
            elif entry['price']['bulk_quote_units'] == 'tonnes':
                d['Bulk Quote Units'].append('tonne')
            else:
                d['Bulk Quote Units'].append(entry['price']['bulk_quote_units'])
        else:
            d['Bulk Quote Units'].append(None)
        if 'lab_scale_units' in entry['price'].keys():
            if entry['price']['lab_scale_units'] == 'tons':
                d['Lab Units'].append('ton')
            elif entry['price']['lab_scale_units'] == 'tonnes':
                d['Lab Units'].append('tonne')
            else:
                d['Lab Units'].append(entry['price']['lab_scale_units'])
        else:
            d['Lab Units'].append(None)
        if entry['price']['type'] == 'lab-scale-log-fit':
            d['Lab-Scale Log Fit?'].append('Lab')
            
        elif entry['price']['type'] == 'quote':
            d['Lab-Scale Log Fit?'].append('Bulk')
        if 'lab-scale-log-fit' in entry['price'].keys():
            if entry['price']['lab-scale-log-fit'] != {}:
                n_pts = len(entry['price']['lab-scale-log-fit']['values'])
                for i in range(0, n_pts):
                    d['Lab Quantity %s' %(i+1)].append(entry['price']['lab-scale-log-fit']['values'][i]['quantity'])
                    d['Lab Price %s' %(i+1)].append(entry['price']['lab-scale-log-fit']['values'][i]['price'])
                if n_pts < 4:
                    for j in range(n_pts, 4):
                        d['Lab Quantity %s' %(j+1)].append(None)
                        d['Lab Price %s' %(j+1)].append(None)
            else:
                for i in range(0,4):
                    d['Lab Quantity %s' %(i+1)].append(None)
                    d['Lab Price %s' %(i+1)].append(None)
                #d['Lab Units'].append(entry['price']['units'])
        if 'quote' in entry['price'].keys():
            if entry['price']['quote']['price']:
                d['Bulk Quote Price ($)'].append(entry['price']['quote']['price'])
            else:
                d['Bulk Quote Price ($)'].append(None)
            if entry['price']['quote']['quantity']:
                d['Bulk Quote Quantity'].append(entry['price']['quote']['quantity'])
            else:
                d['Bulk Quote Quantity'].append(None)
            d['Quote Source'].append(entry['price']['quote']['source'])
            d['Quote Access Date'].append(entry['price']['quote']['date'])
            
            if 'lab-scale-log-fit' not in entry['price'].keys():
                for i in range(0,4):
                    d['Lab Quantity %s' %(i+1)].append(None)
                    d['Lab Price %s' %(i+1)].append(None)
                    
        d['Quote Year'].append(int(entry['price']['year']))

        
    df = pd.DataFrame(data=d)
    df = df[['Material Name', 'Material Type', 'MW (g/mol)', 'Density (g/mL)',
            'Concentration (%)', 'Lab-Scale Log Fit?', 'Lab Quantity 1',
            'Lab Quantity 2', 'Lab Quantity 3', 'Lab Quantity 4', 'Lab Units',
            'Lab Price 1', 'Lab Price 2', 'Lab Price 3', 'Lab Price 4', 
            'Bulk Quote Price ($)', 'Bulk Quote Quantity', 'Bulk Quote Units',
            'Quote Year', 'Quote Source', 'Quote Access Date']]
    df.to_excel(excel_path)
#    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
#        book = load_workbook(excel_path)
#        writer.book = book
#        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#        df.to_excel(writer, sheet_name='Materials Library', header=False, index=False,
#                    startrow=924)
#        writer.save()
#        writer.close()
    
#The commented out sequence above works to add a dataframe in the desired position 
#on excel sheets other than the CCM. Hypothesis is that working with it in the 
#excel table format is throwing something off. Want to repeat my attempt with 
#the equipment library to confirm if it's the table or something else about the
#CCM that's actually causing the problems. Writing to the equipment library has also not worked.
    return df
    
def create_mat_id_dict(excel_path):
    mat_lib_df = get_materials_lib(excel_path)
    mat_id_dict = {}
    mat_lib_df = mat_lib_df.drop(["Bulk Quote Price ($)","Bulk Quote Quantity",'Bulk Quote Units',
                    'Bulk quote quantity in model units','Lab Forecast Unit Price (Output Mass Units)',
                    'Lab Log-Log Intercept','Lab Log-Log Slope','Lab Price 1', 
                    'Lab Price 2','Lab Price 3','Lab Price 4','Lab Q Model Units 1',
                    'Lab Q Model Units 2','Lab Q Model Units 3','Lab Q Model Units 4',
                    'Lab Quantity 1','Lab Quantity 2','Lab Quantity 3', 'Lab Quantity 4',
                    'Lab Quote Count','Lab Unit P 1','Lab Unit P 2','Lab Unit P 3',
                    'Lab Unit P 4','Lab Units','Lab-Scale Log Fit?','Quote Access Date',
                    'Quote Source','Quote Year','Unit price','Unit price dimensions',
                    'Unit price in quote year',"Notes","Basis Cell"],1)
    mat_lib_dict = mat_lib_df.to_dict('records')
    for entry in mat_lib_dict:
        mat_id_dict[entry['Material Name']] = gen_id()
    
    return mat_id_dict

    


"""
The Following section is used to import the equipment library
"""


def get_equipment(excel_path):
    """
    gets the existing materials library as a pandas dataframe. Will be used to create json file
    """
    
    equip_df = pd.read_excel(excel_path,
                                   sheet_name="Equip. Library", skip_footer=1)
    return equip_df

def split_nc(row):
    my_value = row['n / c']
    return my_value

def rename_func_type(row):
    curr_function_type = row['Function Type']
    if curr_function_type == '1 - Power Law':
        new_function_type = "Power Law"
    elif curr_function_type == '2 - Poly Exp':
        new_function_type = "Poly Exp"
    return new_function_type

def equip_to_json(excel_path, json_path, complete=False, version="1.0.0"):
    
    if complete:
        equip_df = get_equipment(excel_path)
    else:
        equip_df = pd.read_excel(excel_path, 
                                   sheet_name="Equip. Library",
                                   skip_footer=1)
    equip_df = equip_df.rename(columns={"Category (not in use)":"category",
                                        "Name":"name", "Year":"year",
                                        "Units for Size, S":"size_unit",
                                        "S lower":"size_min", "S upper":"size_max",
                                        "BM Factor (not in use)":"bm_factor",
                                        "Installation Factor (Garrett)":"installation_factor",
                                        "Note":"note", "Source":"source",
                                        "CEPCI":"cepci","NF Refinery":"nf_refinery",
                                        "Labor Factor":"labor_factor"})
    equip_df = equip_df[equip_df.size_unit.notnull()]
    nrows = len(equip_df)
    equip_df["version"] = [version] * nrows
    #populates version as 1.0.0 for all materials. may become deprecated
    equip_df['c'] = equip_df.apply(lambda row: split_nc(row), axis=1)
    equip_df['n'] = equip_df.apply(lambda row: split_nc(row), axis=1)
    equip_df = equip_df.drop(['n / c'],1)
    equip_df['function_type'] = equip_df.apply(lambda row: rename_func_type(row), axis=1)
    equip_df = equip_df.drop(['Function Type'],1)
    clean_up_lst = ['Pricing Basis Material','Material 1','Material 2',
                    'Material 3', 'Material 4', 'Material 5', 'Material 6',
                    'Material 7', 'Material 8', 'Material 9', 'Material 10',
                    'Factor 1', 'Factor 2', 'Factor 3', 'Factor 4', 'Factor 5',
                    'Factor 6', 'Factor 7', 'Factor 8', 'Factor 9', 'Factor 10']
    #equip_df = equip_df.drop(clean_up_lst,1)
    equip_dict = equip_df.to_dict('records')
    equip_id_dict = get_ids('equip_id_dict')
    for entry in equip_dict:
        if type(entry['year']) == float:
            entry['year'] = str(int(math.floor(entry['year'])))
        if type(entry['note']) == float:
            if math.isnan(entry['note']):
                entry['note'] = None
        if type(entry['bm_factor']) == str:
            entry['bm_factor'] = None
        #convert to null for final version, ask about "use compressor?" for power-recovery turbine
        for key in entry:
            if type(entry[key]) == float:
                if math.isnan(entry[key]): 
                    entry[key] = None
#                elif entry[key] < 0:
#                    entry[key] = -entry[key]
                    #change back to none for final value
        entry['updatedOn'] = int(np.floor(time.time()))
        try:
            entry['id'] = equip_id_dict[entry['name']]
        except KeyError:
            entry['id'] = add_id('equip_id_dict',entry['name'])
        pricing_basis_lst = make_pricing_basis_lst(entry)
        if pricing_basis_lst:
            entry['pricing_basis_materials'] = pricing_basis_lst
        for item in clean_up_lst:
            entry.pop(item, None)
    equip_json = json.dumps(equip_dict,indent=2)
    if json_path != None:
        with open(json_path,'w') as equip_json_out:
            equip_json_out.write(equip_json)
    return equip_json
    
def make_pricing_basis_lst(entry):
    pricing_basis_lst = []
    for idx in range(1,11):
        mat_num = 'Material %s' %idx
        fact_num = 'Factor %s' %idx
        if type(entry[mat_num]) == int:
            continue
        basis_mat_dict = {}
        basis_mat_dict["name"] = entry[mat_num]
        basis_mat_dict["factor"] = entry[fact_num]
        if basis_mat_dict["factor"] != None:
            pricing_basis_lst.append(basis_mat_dict)
    return pricing_basis_lst

def create_equip_id_dict():
    equip_lib_df = get_equipment()
    equip_id_dict = {}
    equip_lib_dict = equip_lib_df.to_dict('records')
    for entry in equip_lib_dict:
        equip_id_dict[entry['Name']] = gen_id()
        
    return equip_id_dict

def equip_json_to_excel(json_path, excel_path):
    with open(json_path,'r') as f:
        equip_json = f.read()
        equip_lst = json.loads(equip_json)
    d={'Category (not in use)':[], 'Name':[], 'Units for Size, S':[], 'S lower':[],
       'S upper':[], 'a':[], 'b':[], 'n/c':[], 'd':[], 'e':[], 'Function Type':[],
       'Source':[], 'CEPCI':[], 'NF Refinery':[], 'Year':[], 'Pricing Basis Material':[],
       'Material 1':[], 'Material 2':[], 'Material 3':[], 'Material 4':[], 
       'Material 5':[], 'Material 6':[], 'Material 7':[], 'Material 8':[], 
       'Material 9':[], 'Material 10':[], 'Factor 1':[], 'Factor 2':[],
       'Factor 3':[], 'Factor 4':[], 'Factor 5':[], 'Factor 6':[], 'Factor 7':[],
       'Factor 8':[], 'Factor 9':[], 'Factor 10':[], 'BM Factor (not in use)':[],
       'Installation Factor (Garrett)':[], 'Labor Factor':[], 'Note':[]}
    for entry in equip_lst:
        d['Category (not in use)'].append(entry['category'])
        d['Name'].append(entry['name'])
        d['Units for Size, S'].append(entry['size_unit'])
        d['S lower'].append(entry['size_min'])
        d['S upper'].append(entry['size_max'])
        d['a'].append(entry['a'])
        d['b'].append(entry['b'])
        d['n/c'].append(entry['c'])
        d['d'].append(entry['d'])
        d['e'].append(entry['e'])
        d['Function Type'].append(entry['function_type'])
        d['Source'].append(entry['source'])
        d['CEPCI'].append(entry['cepci'])
        d['NF Refinery'].append(entry['nf_refinery'])
        d['Year'].append(entry['year'])
        d['Pricing Basis Material'].append(None)
        n_pts = len(entry['pricing_basis_materials'])
        for i in range(0, n_pts):
            d['Material %s' %(i+1)].append(entry['pricing_basis_materials'][i]['name'])
            d['Factor %s' %(i+1)].append(entry['pricing_basis_materials'][i]['factor'])
        if n_pts < 10:
            for j in range(n_pts, 10):
                d['Material %s' %(j+1)].append(None)
                d['Factor %s' %(j+1)].append(None)
        d['BM Factor (not in use)'].append(entry['bm_factor'])
        d['Installation Factor (Garrett)'].append(entry['installation_factor'])
        d['Labor Factor'].append(entry['labor_factor'])
        d['Note'].append(entry['note'])
    
    df=pd.DataFrame(data=d)
    df = df[['Category (not in use)', 'Name', 'Units for Size, S', 'S lower',
   'S upper', 'a', 'b', 'n/c', 'd', 'e', 'Function Type',
   'Source', 'CEPCI', 'NF Refinery', 'Year', 'Pricing Basis Material',
   'Material 1', 'Material 2', 'Material 3', 'Material 4', 
   'Material 5', 'Material 6', 'Material 7', 'Material 8', 
   'Material 9', 'Material 10', 'Factor 1', 'Factor 2',
   'Factor 3', 'Factor 4', 'Factor 5', 'Factor 6', 'Factor 7',
   'Factor 8', 'Factor 9', 'Factor 10', 'BM Factor (not in use)',
   'Installation Factor (Garrett)', 'Labor Factor', 'Note']]
    df.to_excel(excel_path)
    
    return df
        

"""
The section below is designed to populate the spent catalyst library
"""

def make_spent_cat_tables(excel_path):
    with open_workbook(excel_path) as wb:
        sheets = wb.sheets()
        for sheet in sheets:
            if sheet.name == 'Spent Cat Library':
                spent_cat = sheet
                break
        
        for rownum in range(spent_cat.nrows):
            row_values = spent_cat.row_values(rownum)
            if row_values[0] == 'Table 1 - Support-specific data':
                support_start = rownum + 1
            elif row_values[0] == 'Table 2 - Metal-specific data':
                metal_start = rownum + 1
            elif row_values[0] == 'Table 3 - Landfill Fees and Sale Values':
                hazard_start = rownum + 1
            elif row_values[0] == 'Table 4 - Catalyst Bulk Densities (or specified by user)':
                catalyst_start = rownum + 1
            elif row_values[0] == 'Table 5 - RCRA metals and Toxicity Characteristic Leaching Prodedure Limits. Estimated minimum concentrations of RCRA metals as wt. % in catalyst that would pass the TCLP threshold are provided; no warranty is given as to the accuracy or usefulness of these estimates.':
                TCLP_start = rownum + 1
                
#    spent_cat_df = pd.read_excel(excel_path,
#                                   sheet_name="Spent Cat Library",skiprows=3)
        support_end = spent_cat.nrows - (metal_start - 2)
        metal_end = spent_cat.nrows - (hazard_start - 2)
        hazard_end = spent_cat.nrows - (catalyst_start - 2)
        catalyst_end = spent_cat.nrows - (TCLP_start - 2)
    support_df = pd.read_excel(excel_path, sheet_name="Spent Cat Library",
                               skiprows=support_start, usecols='A:V',skipfooter=support_end)
    metal_df = pd.read_excel(excel_path, sheet_name="Spent Cat Library",
                             skiprows=metal_start,usecols='A:N',skipfooter=metal_end)
    hazard_df = pd.read_excel(excel_path,sheet_name="Spent Cat Library",
                              skiprows=hazard_start,usecols='A:H',skipfooter=hazard_end)
    catalyst_df = pd.read_excel(excel_path, sheet_name="Spent Cat Library",
                                skiprows=catalyst_start,usecols='A:D',skipfooter=catalyst_end)
    TCLP_df = pd.read_excel(excel_path, sheet_name="Spent Cat Library",
                            skiprows=TCLP_start, usecols='A:C')
    cleaned_tables = [support_df,metal_df,hazard_df,catalyst_df,TCLP_df]
    return cleaned_tables

def label_cuts(row):
    if type(row['Unnamed: 0']) == float:
        cut_marker = 'blank'
    elif 'Table' in row['Unnamed: 0']:
        cut_marker = 'title'
    else:
        cut_marker = 'data'
    return cut_marker

def separate_tables(df):
    table_starts = []
    table_ends = []
    for row in df.index:
        if df['cut_marker'][row] == 'title':
            table_start = row + 1
            table_starts.append(table_start)
        elif df['cut_marker'][row] == 'blank':
            table_end = row
            table_ends.append(table_end)
    table_ends.append(len(df.index))
    
    table_lst = []
    
    for i in range(0,5):
        table = df[table_starts[i]:table_ends[i]]
        table_lst.append(table)
    
    cleaned_tables = []
    
    for table in table_lst:
        clean_up_lst = []
        for column in table.columns:
            column_values = []
            for idx in table.index:
                if type(table[column][idx]) == float:
                    if math.isnan(table[column][idx]):
                        column_values.append(True)
                    else:
                        column_values.append(False)
                else:
                    column_values.append(False)
            if all(column_values):
                clean_up_lst.append(column)
        for item in clean_up_lst:
            table = table.drop(columns=item, axis=1)
        table = table.drop(columns='cut_marker', axis=1)
        for column in table.columns:
            table = table.rename(columns={column:table[column][min(table.index)]})
        table = table.drop(index=min(table.index))
        cleaned_tables.append(table)
    return cleaned_tables
            
def make_support_dict(cleaned_tables, version='1.0.0'):
    support_df = cleaned_tables[0]
    support_df = support_df.rename(columns={'Support':'name', 
                                            'Loss of catalyst solids in use, fixed, %':'loss_of_catalyst_solids_fixed',
                                            'Loss of metal in use, fixed, %':'loss_of_metal_fixed',
                                            'Loss of catalyst solids in use, slurry/fluidized, %':'loss_of_catalyst_solids_slurry',
                                            'Loss of metal in use, slurry/fluidized, %':'loss_of_metal_slurry'})
    clean_up_lst = ['Incoming fee, ave, $/ft3', 'Incoming fee, high, $/ft3', 
                    'Incoming fee, low, $/ft3', 'Thermal ox. fee, ave, $/lb',
                    'Thermal ox. fee, high, $/lb', 'Thermal ox. fee, low, $/lb',
                    'Metal contaminant fee, ave, $/ft3', 'Metal contaminant fee, high, $/ft3',
                    'Metal contaminant fee, low, $/ft3', 'Fixed bed: Loss of support in use, low, %',
                    'Fixed bed: Loss of support in use, high, %',
                    'Fixed bed: Loss of support in use, ave, %',
                    'Fixed bed: Loss of metal in use, low, %',
                    'Fixed bed: Loss of metal in use, high, %',
                    'Fixed bed: Loss of metal in use, ave, %',
                    'Slurry/fluidized bed: Loss of support in use, low, %',
                    'Slurry/fluidized bed: Loss of support in use, high, %', 
                    'Slurry/fluidized bed: Loss of support in use, ave, %',
                    'Slurry/fluidized bed: Loss of metal in use, low, %',
                    'Slurry/fluidized bed: Loss of metal in use, high, %',
                    'Slurry/fluidized bed: Loss of metal in use, ave, %']
    support_dict = support_df.to_dict('records')
    support_id_dict = get_ids('support_id_dict')
    for entry in support_dict:
        entry['incoming_fee'] = make_incoming_dict(entry)
        entry['thermal_ox_fee'] = make_thermal_ox_dict(entry)
        entry['metal_contaminant_fee'] = make_metal_contaminant_dict(entry)
        entry['loss_of_catalyst_solids_fixed'] = entry['Fixed bed: Loss of support in use, ave, %']
        entry['loss_of_metal_fixed'] = entry['Fixed bed: Loss of metal in use, ave, %']
        entry['loss_of_catalyst_solids_slurry'] = entry['Slurry/fluidized bed: Loss of support in use, ave, %']
        entry['loss_of_metal_slurry'] = entry['Slurry/fluidized bed: Loss of metal in use, ave, %']
#        entry['loss_of_catalyst_solids_fixed'] = make_fixed_support_loss_dict(entry)
#        entry['loss_of_metal_fixed'] = make_fixed_metal_loss_dict(entry)
#        entry['loss_of_catalyst_solids_slurry'] = make_slurry_support_loss_dict(entry)
#        entry['loss_of_metal_slurry'] = make_slurry_support_loss_dict(entry)
        entry['version'] = version
        entry['updatedOn'] = int(np.floor(time.time()))
        try:
            entry['id'] = support_id_dict[entry['name']]
        except KeyError:
            entry['id'] = add_id('support_id_dict',entry['name'])
        for key in entry.keys():
            if type(entry[key]) == float:
                if math.isnan(entry[key]):
                    entry[key] = None
        for item in clean_up_lst:
            entry.pop(item, None)
    return support_dict

def create_support_id_dict(cleaned_tables):
    support_df = cleaned_tables[0]
    support_df = support_df.rename(columns={'Support':'name', 
                                            'Loss of catalyst solids in use, fixed, %':'loss_of_catalyst_solids_fixed',
                                            'Loss of metal in use, fixed, %':'loss_of_metal_fixed',
                                            'Loss of catalyst solids in use, slurry/fluidized, %':'loss_of_catalyst_solids_slurry',
                                            'Loss of metal in use, slurry/fluidized, %':'loss_of_metal_slurry'})
    support_dict = support_df.to_dict('records')
    support_id_dict = {}
    for entry in support_dict:
        support_id_dict[entry['name']] = gen_id() 
        
    return support_id_dict

def make_fixed_support_loss_dict(entry):
    fixed_support_loss_dict = {}
    fixed_support_loss_dict['baseline'] = entry['Fixed bed: Loss of support in use, ave, %']
    fixed_support_loss_dict['high'] = entry['Fixed bed: Loss of support in use, high, %']
    fixed_support_loss_dict['low'] = entry['Fixed bed: Loss of support in use, low, %']
    for key in fixed_support_loss_dict.keys():
        if type(fixed_support_loss_dict[key]) == float:
            if math.isnan(fixed_support_loss_dict[key]):
                fixed_support_loss_dict[key] = None
    return fixed_support_loss_dict

def make_fixed_metal_loss_dict(entry):
    fixed_metal_loss_dict = {}
    fixed_metal_loss_dict['baseline'] = entry['Fixed bed: Loss of metal in use, ave, %']
    fixed_metal_loss_dict['high'] = entry['Fixed bed: Loss of metal in use, high, %']
    fixed_metal_loss_dict['low'] = entry['Fixed bed: Loss of metal in use, low, %']
    for key in fixed_metal_loss_dict.keys():
        if type(fixed_metal_loss_dict[key]) == float:
            if math.isnan(fixed_metal_loss_dict[key]):
                fixed_metal_loss_dict[key] = None
    return fixed_metal_loss_dict

def make_slurry_support_loss_dict(entry):
    slurry_support_loss_dict = {}
    slurry_support_loss_dict['baseline'] = entry['Slurry/fluidized bed: Loss of support in use, ave, %']
    slurry_support_loss_dict['high'] = entry['Slurry/fluidized bed: Loss of support in use, high, %']
    slurry_support_loss_dict['low'] = entry['Slurry/fluidized bed: Loss of support in use, low, %']
    for key in slurry_support_loss_dict.keys():
        if type(slurry_support_loss_dict[key]) == float:
            if math.isnan(slurry_support_loss_dict[key]):
                slurry_support_loss_dict[key] = None
    return slurry_support_loss_dict

def make_slurry_metal_loss_dict(entry):
    slurry_metal_loss_dict = {}
    slurry_metal_loss_dict['baseline'] = entry['Slurry/fluidized bed: Loss of metal in use, ave, %']
    slurry_metal_loss_dict['high'] = entry['Slurry/fluidized bed: Loss of metal in use, high, %']
    slurry_metal_loss_dict['low'] = entry['Slurry/fluidized bed: Loss of metal in use, low, %']
    for key in slurry_metal_loss_dict.keys():
        if type(slurry_metal_loss_dict[key]) == float:
            if math.isnan(slurry_metal_loss_dict[key]):
                slurry_metal_loss_dict[key] = None
    return slurry_metal_loss_dict

def make_incoming_dict(entry):
    incoming_dict = {}
    incoming_dict['baseline'] = entry['Incoming fee, ave, $/ft3']
    incoming_dict['high'] = entry['Incoming fee, high, $/ft3']
    incoming_dict['low'] = entry['Incoming fee, low, $/ft3']
    return incoming_dict

def make_thermal_ox_dict(entry):
    thermal_ox_dict = {}
    thermal_ox_dict['baseline'] = entry['Thermal ox. fee, ave, $/lb']
    thermal_ox_dict['high'] = entry['Thermal ox. fee, high, $/lb']
    thermal_ox_dict['low'] = entry['Thermal ox. fee, low, $/lb']
    return thermal_ox_dict

def make_metal_contaminant_dict(entry):
    metal_contaminant_dict = {}
    metal_contaminant_dict['baseline'] = entry['Metal contaminant fee, ave, $/ft3']
    metal_contaminant_dict['high'] = entry['Metal contaminant fee, high, $/ft3']
    metal_contaminant_dict['low'] = entry['Metal contaminant fee, low, $/ft3']
    return metal_contaminant_dict

def make_metal_dict(cleaned_tables, version='1.0.0'):
    metal_df = cleaned_tables[1]
    metal_df = metal_df.rename(columns={'Metal':'name', 
                                        'Refining charge, $/troy oz recovered':'refining_charge',
                                        'Note':'note', 'PGM/Noble (Refining charge yes/no)':'has_refining_charge',
                                        'Spot Price ($)':'spot_price','Unit':'unit',
                                        'Year':'year','Source':'source'})
    metal_dict = metal_df.to_dict('records')
    clean_up_lst = ['Price $ / Model Mass Unit','Price Scaled to Basis Year','Units',
                    'Loss during refining, low, %','Loss during refining, high, %','Loss during refining, ave, %']
    metal_id_dict = get_ids('metal_id_dict')
    for entry in metal_dict:
        entry['loss'] = make_metal_loss_dict(entry)
        for item in clean_up_lst:
            entry.pop(item, None)
        entry['has_refining_charge'] = entry['has_refining_charge'].lower()
        entry['year'] = str(entry['year'])
        if entry['unit'] == 'tonne': #correct this when the webtool is updated to take tonne
            entry['unit'] = 'tonnes'
        if entry['unit'] == 'ton':
            entry['unit'] = 'tons'
            #entry['unit'] = 'kg'
            #entry['spot_price'] = entry['spot_price']*0.00110231
#        if type(entry['note']) == float:
#            if math.isnan(entry['note']):
#                entry['note'] = None
#        if type(entry['refining_charge']) == float:
#            if math.isnan(entry['refining_charge']):
#                entry['refining_charge'] = None
        for key in entry:
            if type(entry[key]) == float:
                if math.isnan(entry[key]):
                    entry[key] = None
#use above block during final version
        entry['version'] = version
        try:
            entry['id'] = metal_id_dict[entry['name']]
        except KeyError:
            entry['id'] = add_id('metal_id_dict',entry['name'])
        entry['updatedOn'] = int(np.floor(time.time()))
    return metal_dict

def create_metal_id_dict(cleaned_tables):
    metal_df = cleaned_tables[1]
    metal_df = metal_df.rename(columns={'Metal':'name', 
                                        'Refining charge, $/troy oz recovered':'refining_charge',
                                        'Note':'note', 'PGM/Noble (Refining charge yes/no)':'has_refining_charge',
                                        'Spot Price ($)':'spot_price','Unit':'unit',
                                        'Year':'year','Source':'source'})
    metal_dict = metal_df.to_dict('records')
    metal_id_dict = {}
    for entry in metal_dict:
        metal_id_dict[entry['name']] = gen_id()
        
    return metal_id_dict

def make_metal_loss_dict(entry):
    metal_loss_dict = {}
    metal_loss_dict['baseline'] = entry['Loss during refining, ave, %']
    metal_loss_dict['high'] = entry['Loss during refining, high, %']
    metal_loss_dict['low'] = entry['Loss during refining, low, %']
    if type(metal_loss_dict['high']) == float:
        if math.isnan(metal_loss_dict['high']):
            metal_loss_dict['high'] = None
    if type(metal_loss_dict['low']) == float:
        if math.isnan(metal_loss_dict['low']):
            metal_loss_dict['low'] = None
    return metal_loss_dict

def make_hazard_dict(cleaned_tables, version="1.0.0"):
    hazard_df = cleaned_tables[2]
    hazard_df = hazard_df.rename(columns={'Catalyst Hazard Class':'name','Note':'note'})
    hazard_dict = hazard_df.to_dict('records')
    clean_up_lst = ['Landfill fee, low, $/lb','Landfill fee, high, $/lb',
                    'Landfill fee, ave, $/lb','Sale value, low, $/lb',
                    'Sale value, high, $/lb','Sale value, ave, $/lb']
    hazard_id_dict = get_ids('hazard_id_dict')
    for entry in hazard_dict:
        try:
            entry['id'] = hazard_id_dict[entry['name']]
        except KeyError:
            entry['id'] = add_id('hazard_id_dict',entry['name'])
        entry['updatedOn'] = int(np.floor(time.time()))
        entry['version'] = version
        entry['landfill_fee'] = make_landfill_dict(entry)
        entry['sale_value'] = make_sale_dict(entry)
        if type(entry['note']) == float:
            if math.isnan(entry['note']):
                entry['note'] = None
        for item in clean_up_lst:
            entry.pop(item, None)
    return hazard_dict

def create_hazard_id_dict(cleaned_tables):
    hazard_df = cleaned_tables[2]
    hazard_df = hazard_df.rename(columns={'Catalyst Hazard Class':'name','Note':'note'})
    hazard_dict = hazard_df.to_dict('records')
    hazard_id_dict = {}
    for entry in hazard_dict:
        hazard_id_dict[entry['name']] = gen_id()
        
    return hazard_id_dict

def make_landfill_dict(entry):
    landfill_dict = {}
    landfill_dict['baseline'] = entry['Landfill fee, ave, $/lb']
    landfill_dict['low'] = entry['Landfill fee, low, $/lb']
    landfill_dict['high'] = entry['Landfill fee, high, $/lb']
    if math.isnan(landfill_dict['low']):
        landfill_dict['low'] = None
    if math.isnan(landfill_dict['high']):
        landfill_dict['high'] = None
    return landfill_dict

def make_sale_dict(entry):
    sale_dict = {}
    sale_dict['baseline'] = entry['Sale value, ave, $/lb']
    sale_dict['low'] = entry['Sale value, low, $/lb']
    sale_dict['high'] = entry['Sale value, high, $/lb']
    if math.isnan(sale_dict['baseline']):
        sale_dict['baseline'] = None
    if math.isnan(sale_dict['low']):
        sale_dict['low'] = None
    if math.isnan(sale_dict['high']):
       sale_dict['high'] = None
    return sale_dict

def make_density_dict(cleaned_tables,version="1.0.0"):
    density_df = cleaned_tables[3]
    density_df = density_df.rename(columns={"Catalyst":"name", "ρ (lb/ft3)":"density",
                                            "Note":"note"})
                                            #"ρ (kg/m3)":"density_si","Note":"note"})
    density_df.drop("ρ (kg/m3)",axis=1,inplace=True)
    density_dict = density_df.to_dict('records')
    density_id_dict = get_ids('density_id_dict')
    for entry in density_dict:
        try:
            entry['id'] = density_id_dict[entry['name']]
        except KeyError:
            entry['id'] = add_id('density_id_dict',entry['name'])
        entry['version'] = version
        entry['updatedOn'] = int(np.floor(time.time()))
        entry['density_unit'] = 'lb/ft3'
    return density_dict

def create_density_id_dict(cleaned_tables):
    density_df = cleaned_tables[3]
    density_df = density_df.rename(columns={"Catalyst":"name", "ρ (lb/ft3)":"density",
                                            "Note":"note"})
                                            #"ρ (kg/m3)":"density_si","Note":"note"})
    density_dict = density_df.to_dict('records')
    density_id_dict = {}
    for entry in density_dict:
        density_id_dict[entry['name']] = gen_id()

    
    return density_id_dict

def spent_cat_to_json(excel_path, json_path):
#    spent_cat_df = make_spent_cat_tables(excel_path)
#    spent_cat_df['cut_marker'] = spent_cat_df.apply(lambda row: label_cuts(row), axis=1)
    cleaned_tables = make_spent_cat_tables(excel_path)
    spent_cat_dict = {}
    spent_cat_dict['spent_cat_support'] = make_support_dict(cleaned_tables)
    spent_cat_dict['spent_cat_metal'] = make_metal_dict(cleaned_tables)
    spent_cat_dict['spent_cat_hazard'] = make_hazard_dict(cleaned_tables)
    spent_cat_dict['spent_cat_bulk_density'] = make_density_dict(cleaned_tables)
    spent_cat_json = json.dumps(spent_cat_dict,indent=2)
    if json_path != None:
        with open(json_path,'w') as spent_cat_out:
            spent_cat_out.write(spent_cat_json)
    return spent_cat_json


"""
The following section is designed to write catalyst estimates from the excel tool
to json format for upload to the webtool
"""
def estimate_to_json(excel_path, json_path, version="1.0.0"):
    with open_workbook(excel_path) as wb:
        sheets = wb.sheets()
        for sheet in sheets:
            if sheet.name == '1 Inputs':
                inputs = sheet
            elif sheet.name == '2 Materials':
                materials = sheet 
            elif sheet.name == '3b Equip':
                equip = sheet
            elif sheet.name == '3e OpEx':
                opex = sheet
            elif sheet.name == '5 Outputs':
                outputs = sheet
        est_dict = {}
        est_dict['id'] = gen_id()
        est_dict['version'] = version
        est_dict['updatedOn'] = int(np.floor(time.time()))
        for rownum in range(inputs.nrows):
            row_value = inputs.row_values(rownum)
            if "Estimate Name" in row_value:
                tmp_lst = [i for i,x in enumerate(row_value) if x == 'Estimate Name']
                est_dict['name'] = inputs.row_values(rownum+1)[tmp_lst[0]]
            elif 'Basis Year' in row_value:
                tmp_lst = [i for i,x in enumerate(row_value) if x == 'Basis Year']
                try:
                    est_dict['basis_year'] = str(int(row_value[tmp_lst[0] + 1]))
                except ValueError:
                    pass
            elif 'Currency' in row_value:
                tmp_lst = [i for i,x in enumerate(row_value) if x == 'Currency']
                if row_value[tmp_lst[0] + 1] == 'USD ($)':
                    est_dict['currency'] = 'USD, $'
            elif 'Mass Unit' in row_value:
                tmp_lst = [i for i,x in enumerate(row_value) if x == 'Mass Unit']
                est_dict['mass_unit'] = row_value[tmp_lst[0] + 1]
                basis_unit = est_dict['mass_unit']
            elif 'Design Production, Annual' in row_value:
                est_dict['design_production'] = locate_data(row_value, 'Design Production, Annual')
            elif 'Capacity Factor' in row_value:
                est_dict['capacity_factor'] = locate_data(row_value, 'Capacity Factor')
            elif 'Operating Hours (Labor)' in row_value:
                est_dict['operating_hours_labor'] = locate_data(row_value, 'Operating Hours (Labor)')
            elif 'Return on capital invested (pre-tax)' in row_value:
                est_dict['ROI'] = locate_data(row_value,'Return on capital invested (pre-tax)')
            elif 'On-Stream Factor' in row_value:
                est_dict['stream_factor'] = locate_data(row_value, 'On-Stream Factor')
                
        est_equip_lst, equip_mass_unit, equip_time_unit, catalyst_or_AP, reference_design_production  = make_est_equip_lst(excel_path, est_dict['id'], version)
        est_dict['equip_mass_unit'] = equip_mass_unit
        est_dict['equip_time_unit'] = equip_time_unit
        est_dict['catalyst_or_AP'] = catalyst_or_AP
        est_dict['reference_design_production'] = reference_design_production
        est_dict['estimate_equipment'] = est_equip_lst
        est_mat_lst = make_est_mat_lst(excel_path, est_dict['id'], version)
        est_dict['estimate_materials'] = est_mat_lst
        spent_cat_metal_dict, spent_cat_support_dict, spent_cat_hazard_dict, spent_cat_density_dict = make_est_spent_cat(excel_path,est_dict['id'],version)
        est_dict['estimate_spent_cat_metal'] = spent_cat_metal_dict
        est_dict['estimate_spent_cat_hazard'] = spent_cat_hazard_dict
        est_dict['estimate_spent_cat_support'] = spent_cat_support_dict
        est_dict['estimate_spent_cat_bulk_density'] = spent_cat_density_dict
        utility_lst = make_est_process_utilities(excel_path, est_dict['id'], version, basis_unit)
        est_dict['process_utilities'] = utility_lst
        est_dict['factored_capital_costs'] = make_est_cap_ex(excel_path, est_dict['id'], version)
        est_dict['factored_operating_costs'] = make_est_op_ex(excel_path, est_dict['id'], version)
        
#        for rownum in range(equip.nrows):
#            row_value = equip.row_values(rownum)
#            if ' < select mass unit' in row_value:     
#                tmp_lst = [i for i,x in enumerate(row_value) if x == ' < select mass unit']
#                est_dict['equip_mass_unit'] = row_value[tmp_lst[0] - 1]
#            elif ' < select time unit' in row_value:
#                tmp_lst = [i for i,x in enumerate(row_value) if x == ' < select time unit']
#                equip_time_unit = row_value[tmp_lst[0] - 1]
#                if equip_time_unit == 'hour':
#                    est_dict['equip_time_unit'] = 'hr'
#                elif equip_time_unit == 'week':
#                    est_dict['equip_time_unit'] = 'wk'
#                else:
#                    est_dict['equip_time_unit'] = equip_time_unit
#            elif ' < select catalyst or AP' in row_value:
#                tmp_lst = [i for i,x in enumerate(row_value) if x == ' < select catalyst or AP']
#                est_dict['catalyst_or_AP'] = row_value[tmp_lst[0] - 1] #this entry shows up in the schema, but not the sample estimate I pulled from the web tool
#            elif 'Reference Design Production Rate (do not enter a formula)' in row_value:
#                tmp_lst = [i for i,x in enumerate(row_value) if x == 'Reference Design Production Rate (do not enter a formula)']
#                est_dict['reference_design_production'] = row_value[tmp_lst[0] + 1]
                
        for rownum in range(materials.nrows):
            row_value = materials.row_values(rownum)
            if 'Yield Type: % Yield or Mass?' in row_value:
                tmp_lst = [i for i,x in enumerate(row_value) if x == 'Yield Type: % Yield or Mass?']
                est_dict['yield_type'] = row_value[tmp_lst[0] + 1]
            elif ('Finished Catalyst Mass Yield' in row_value) or ('Active Phase Mass Yield' in row_value) or ('% Yield of Finished Catalyst' in row_value):
                tmp_lst = [i for i,x in enumerate(row_value) if (x == 'Finished Catalyst Mass Yield') or (x == 'Active Phase Mass Yield') or (x == '% Yield of Finished Catalyst')]
                est_dict['yield'] = {}
                est_dict['yield']['baseline'] = row_value[tmp_lst[0] + 1]
                est_dict['yield']['low'] = row_value[tmp_lst[0] + 2]
                est_dict['yield']['high'] = row_value[tmp_lst[0] + 3]
            elif 'Stoichiometric Ratio AP/metal' in row_value:
                est_dict['stoichiometric_ratio'] = locate_data(row_value, 'Stoichiometric Ratio AP/metal')
            elif 'Active Phase Moleuclar Weight' in row_value:
                est_dict['active_phase_molecular_weight'] = locate_data(row_value, 'Active Phase Molecular Weight')
            elif 'Active Phase Weight Percent' in row_value:
                est_dict['active_phase_weight_percent'] = locate_data(row_value, 'Active Phase Weight Percent')
            elif 'Losses Due to Waste/Spoilage' in row_value:
                est_dict['losses_waste_spoilage'] = locate_data(row_value, 'Losses Due to Waste/Spoilage')
                
        for rownum in range(opex.nrows):
            row_value = opex.row_values(rownum)
            if 'Direct Labor Rate' in row_value:
                est_dict['operating_direct_labor_rate'] = locate_data(row_value, 'Direct Labor Rate', sensitivity=False)
                est_dict['direct_labor_rate'] = locate_data(row_value, 'Direct Labor Rate')
            elif 'Direct Labor Operators (rounded up)' in row_value:
                est_dict['direct_labor_operators'] = locate_data(row_value,'Direct Labor Operators (rounded up)')
                
        for rownum in range(outputs.nrows):
            row_value = outputs.row_values(rownum)
            if "Annual, Monthly, Weekly, Daily Cost?" in row_value:
                tmp_lst = [i for i,x in enumerate(row_value) if x == 'Annual, Monthly, Weekly, Daily Cost?']
                out_time = row_value[tmp_lst[0] + 1]
                if out_time == 'Annual':
                    est_dict['output_time_unit'] = 'year'
                elif out_time == 'Monthly':
                    est_dict['output_time_unit'] = 'month'
                elif out_time == 'Weekly':
                    est_dict['output_time_unit'] = 'week'
                elif out_time == 'Daily':
                    est_dict['output_time_unit'] = 'day'
            elif 'Unit Cost in Cents or Dollars (USD, $)' in row_value:
                tmp_lst = [i for i,x in enumerate(row_value) if x == 'Unit Cost in Cents or Dollars (USD, $)']
                out_currency = row_value[tmp_lst[0] + 1]
                if out_currency == 'Dollars':
                    est_dict['output_currency'] = 'dollars'
                elif out_currency == 'Cents':
                    est_dict['output_currency'] = 'cents'
                    
        for key in est_dict.keys():
            if type(est_dict[key]) == float:
                if math.isnan(est_dict[key]):
                    est_dict[key] = None
            elif type(est_dict[key]) == str:
                if len(est_dict[key]) == 0:
                    est_dict[key] = None
            elif type(est_dict[key]) == dict:
                for key_i in est_dict[key].keys():
                    if type(est_dict[key][key_i]) == float:
                        if math.isnan(est_dict[key][key_i]):
                            est_dict[key][key_i] == None
                    elif type(est_dict[key][key_i]) == str:
                        if len(est_dict[key][key_i]) == 0:
                            est_dict[key][key_i] = None
                    elif type(est_dict[key][key_i]) == dict:
                        for key_ii in est_dict[key][key_i]:
                            if type(est_dict[key][key_i][key_ii]) == float:
                                if math.isnan(est_dict[key][key_i][key_ii]):
                                    est_dict[key][key_i][key_ii] == None
                            elif type(est_dict[key][key_i][key_ii]) == str:
                                if len(est_dict[key][key_i][key_ii]) == 0:
                                    est_dict[key][key_i][key_ii] = None
            elif type(est_dict[key]) == list:
                for idx in range(0, len(est_dict[key])):
                    if type(est_dict[key][idx]) == float:
                        if math.isnan(est_dict[key][idx]):
                            math.isnan[key][idx] = None
                    elif type(est_dict[key][idx]) == str:
                        if len(est_dict[key][idx]) == 0:
                            est_dict[key][idx] = None
                    elif type(est_dict[key][idx]) == dict:
                        for key_ii in est_dict[key][idx].keys():
                            if type(est_dict[key][idx][key_ii]) == float:
                                if math.isnan(est_dict[key][idx][key_ii]):
                                    est_dict[key][idx][key_ii] = None
                            elif type(est_dict[key][idx][key_ii]) == str:
                                if len(est_dict[key][idx][key_ii]) == 0:
                                    est_dict[key][idx][key_ii] = None
                            elif type(est_dict[key][idx][key_ii]) == dict:
                                for key_iii in est_dict[key][idx][key_ii].keys():
                                    if type(est_dict[key][idx][key_ii][key_iii]) == float:
                                        if math.isnan(est_dict[key][idx][key_ii][key_iii]):
                                            est_dict[key][idx][key_ii][key_iii] = None
                                    elif type(est_dict[key][idx][key_ii][key_iii]) == str:
                                        if len(est_dict[key][idx][key_ii][key_iii]) == 0:
                                            est_dict[key][idx][key_ii][key_iii] = None
    
    est_lst = []
    est_lst.append(est_dict)
    est_json = json.dumps(est_lst,indent=2)
    with open(json_path,'w') as est_destination:
        est_destination.write(est_json)
    return est_json

def locate_data(row_value,excel_name,sensitivity=True):
    tmp_lst = [i for i,x in enumerate(row_value) if x == excel_name]
    if sensitivity:
        tmp_dict = {}
        tmp_dict['baseline'] = row_value[tmp_lst[0] + 1]
        if row_value[tmp_lst[0] + 2]:
            tmp_dict['low'] = row_value[tmp_lst[0] + 2]
            if tmp_dict['low'] == 'n/a':
                tmp_dict['low'] = None
        if row_value[tmp_lst[0] + 3]:
            tmp_dict['high'] = row_value[tmp_lst[0] + 3]
            if tmp_dict['high'] == 'n/a':
                tmp_dict['high'] = None
        return tmp_dict
    else:
        output = row_value[tmp_lst[0]+1]
        return output
        
def make_est_equip_lst(excel_path, est_id, version):
    with open_workbook(excel_path) as wb:
        for sheet in wb.sheets():
            if sheet.name == '3b Equip':
                equip = sheet
                break
    user_entry_lims = []
    metal_carbide_lims = []
    nanoparticle_batch_lims = []
    nanoparticle_flow_lims = []
    wet_impregnation_lims = []
    zeolite_lims = []
    for rownum in range(equip.nrows):
        row_value = equip.row_values(rownum)
        nonempty_rowval = [x for x in row_value if x != '']
        if 'Select a Process Template or Choose "Custom Process"' in row_value:
            template_row = equip.row_values(rownum + 1)
            template = [x for x in template_row if x != '']
            try:
                template = template[0]
            except IndexError:
                template = 'Custom Process'
        elif nonempty_rowval:
            if 'User Entry' in nonempty_rowval[0]:
                user_entry_lims.append(rownum)
            elif 'Process Template: Metal Carbide on Metal Oxide' in nonempty_rowval[0]:
                metal_carbide_lims.append(rownum)
            elif 'Process Template: Nanoparticles - Batch Synthesis' in nonempty_rowval[0]:
                nanoparticle_batch_lims.append(rownum)
            elif 'Process Template: Nanoparticles - Flow Synthesis' in nonempty_rowval[0]:
                nanoparticle_flow_lims.append(rownum)
            elif 'Process Template: Wet Impregnation - Metal on Metal Oxide' in nonempty_rowval[0]:
                wet_impregnation_lims.append(rownum)
            elif 'Process Template: Zeolite for FCC' in nonempty_rowval[0]:
                zeolite_lims.append(rownum)
    
    if template == 'Custom Process':
        process_lims = user_entry_lims
    elif template == 'Metal Carbide on Metal Oxide':
        process_lims = metal_carbide_lims
    elif template == 'Nanoparticles - Batch Synthesis':
        process_lims = nanoparticle_batch_lims
    elif template == 'Nanoparticles - Flow Synthesis':
        process_lims = nanoparticle_flow_lims
    elif template == 'Wet Impregnation - Metal on Metal Oxide':
        process_lims = wet_impregnation_lims
    elif template == 'Zeolite for FCC':
        process_lims = zeolite_lims
    
    for rownum in range(process_lims[0], process_lims[1]):
        row_value = equip.row_values(rownum)
        if 'Equipment Type' in row_value:
            df_start = rownum
        elif ' < select mass unit' in row_value:     
            tmp_lst = [i for i,x in enumerate(row_value) if x == ' < select mass unit']
            equip_mass_unit = row_value[tmp_lst[0] - 1]
        elif ' < select time unit' in row_value:
            tmp_lst = [i for i,x in enumerate(row_value) if x == ' < select time unit']
            equip_time_unit = row_value[tmp_lst[0] - 1]
            if equip_time_unit == 'hour':
                equip_time_unit = 'hr'
            elif equip_time_unit == 'week':
                equip_time_unit = 'wk'
            else:
                equip_time_unit = equip_time_unit
        elif ' < select catalyst or AP' in row_value:
            tmp_lst = [i for i,x in enumerate(row_value) if x == ' < select catalyst or AP']
            catalyst_or_AP = row_value[tmp_lst[0] - 1] #this entry shows up in the schema, but not the sample estimate I pulled from the web tool
        elif 'Production Rate for this Process Template' in row_value:
            tmp_lst = [i for i,x in enumerate(row_value) if x == 'Production Rate for this Process Template']
            reference_design_production = row_value[tmp_lst[0] + 1]
            
    df = pd.read_excel(excel_path, sheet_name='3b Equip', skiprows=df_start, usecols='C:H', skipfooter=equip.nrows-process_lims[1])
    df = df.loc[np.isfinite(df['Quantity'])]
    equip_ids = get_ids('equip_id_dict') 
    est_equip_lst = []
    equip_lib = equip_to_json(excel_path, None)
    equip_lib = json.loads(equip_lib)
    for entry in df.iterrows():
        entry = entry[1]
        equip_dict = {}
        equip_dict['id'] = gen_id()
        equip_dict['estimate_id'] = est_id
        equip_dict['equipment_id'] = equip_ids[entry['Equipment Type']]
        equip_dict['construction_material'] = {}
        equip_dict['construction_material']['name'] = entry['Material of Construction']
        for equipment in equip_lib:
            if equipment['id'] == equip_dict['equipment_id']:
                for eq_material in equipment['pricing_basis_materials']:
                    if eq_material['name'] == equip_dict['construction_material']['name']:
                        equip_dict['construction_material']['factor'] = eq_material['factor']
                        break
                equip_dict['equipment'] = equipment
                break
        equip_dict['quantity'] = entry['Quantity']
        equip_dict['size'] = entry['Size']
        equip_dict['updatedOn'] = int(np.floor(time.time()))
        equip_dict['version'] = version
        est_equip_lst.append(equip_dict)
        
    return est_equip_lst, equip_mass_unit, equip_time_unit, catalyst_or_AP, reference_design_production

def make_est_mat_lst(excel_path, est_id, version):
    with open_workbook(excel_path) as wb:
        for sheet in wb.sheets():
            if sheet.name == '2 Materials':
                materials = sheet
                break
    for rownum in range(materials.nrows):
        row_value = materials.row_values(rownum)
        if 'Metal Sources' in row_value:
           metal_source_df_start = rownum + 3
        elif 'Supports' in row_value:
            support_df_start = rownum + 3
        elif 'Other Materials' in row_value:
            other_df_start = rownum + 3
    metal_source_df = pd.read_excel(excel_path, sheet_name='2 Materials', 
                                    skiprows=metal_source_df_start, usecols='C:P')
    end_metal = 100
    for entry in metal_source_df.iterrows():
        if end_metal == 100:
            if not np.isfinite(entry[1]['Quantity (Q)']):
                end_metal = entry[0]
                break
    metal_source_df = metal_source_df[metal_source_df.index < end_metal]
#    metal_source_df = metal_source_df[np.isfinite(metal_source_df['Quantity (Q)'])]
    support_df = pd.read_excel(excel_path, sheet_name='2 Materials',
                               skiprows=support_df_start, usecols='C:P')
    end_support = 100
    for entry in support_df.iterrows():
        if end_support == 100:
            if not np.isfinite(entry[1]['Quantity (Q)']):
                end_support = entry[0]
                break
    support_df = support_df[support_df.index < end_support]
#    support_df = support_df[np.isfinite(support_df['Quantity (Q)'])]
    other_df = pd.read_excel(excel_path, sheet_name='2 Materials',
                             skiprows=other_df_start, usecols='C:P', skipfooter=4)
    end_other = 100
    for entry in other_df.iterrows():
        if end_other == 100:
            if not np.isfinite(entry[1]['Quantity (Q)']):
                end_other = entry[0]
                break
    other_df = other_df[other_df.index < end_other]
#    other_df = other_df[np.isfinite(other_df['Quantity (Q)'])]
    est_mat_lst = []
    mat_lib = materials_to_json(excel_path, None)
    mat_lib = json.loads(mat_lib)
    mat_ids = get_ids('mat_id_dict')
    for entry in metal_source_df.iterrows():
        entry = entry[1]
        metal_dict = {}
        metal_dict['id'] = gen_id()
        metal_dict['estimate_id'] = est_id
        metal_dict['material_id'] = mat_ids[entry['Material Name']]
        metal_dict['version'] = version
        metal_dict['category'] = 'metal'
        metal_dict['name'] = entry['Material Name']
        metal_dict['quantity_unit'] = entry['Unit']
        metal_dict['quantity'] = {}
        metal_dict['quantity']['baseline'] = entry['Quantity (Q)']
        metal_dict['quantity']['high'] = entry['Q high']
        metal_dict['quantity']['low'] = entry['Q low']
        if type(metal_dict['quantity']['high']) == float:
            if math.isnan(metal_dict['quantity']['high']):
                metal_dict['quantity'] = {}
                metal_dict['quantity']['baseline'] = entry['Quantity (Q)']
        metal_dict['material_unit_price'] = {}
        metal_dict['material_unit_price']['low'] = entry['Price low']
        metal_dict['material_unit_price']['high'] = entry['Price high']
        metal_dict['updatedOn'] = int(np.floor(time.time()))
        for mat in mat_lib:
            if mat['id'] == metal_dict['material_id']:
                metal_dict['material'] = mat
                break
        
        est_mat_lst.append(metal_dict)
        
    for entry in support_df.iterrows():
        entry = entry[1]
        support_dict = {}
        support_dict['id'] = gen_id()
        support_dict['estimate_id'] = est_id
        support_dict['material_id'] = mat_ids[entry['Material Name']]
        support_dict['version'] = version
        support_dict['category'] = 'support'
        support_dict['name'] = entry['Material Name']
        support_dict['quantity_unit'] = entry['Unit']
        support_dict['quantity'] = {}
        support_dict['quantity']['baseline'] = entry['Quantity (Q)']
        support_dict['quantity']['high'] = entry['Q high']
        support_dict['quantity']['low'] = entry['Q low']
        if type(support_dict['quantity']['high']) == float:
            if math.isnan(support_dict['quantity']['high']):
                support_dict['quantity'] = {}
                support_dict['quantity']['baseline'] = entry['Quantity (Q)']
        support_dict['material_unit_price'] = {}
        support_dict['material_unit_price']['low'] = entry['Price low']
        support_dict['material_unit_price']['high'] = entry['Price high']
        support_dict['updatedOn'] = int(np.floor(time.time()))
        for mat in mat_lib:
            if mat['id'] == support_dict['material_id']:
                support_dict['material'] = mat
                break
            
        est_mat_lst.append(support_dict)
        
    for entry in other_df.iterrows():
        entry = entry[1]
        other_dict = {}
        other_dict['id'] = gen_id()
        other_dict['estimate_id'] = est_id
        other_dict['material_id'] = mat_ids[entry['Material Name']]
        other_dict['version'] = version
        other_dict['category'] = 'other'
        other_dict['name'] = entry['Material Name']
        other_dict['quantity_unit'] = entry['Unit']
        other_dict['quantity'] = {}
        other_dict['quantity']['baseline'] = entry['Quantity (Q)']
        other_dict['quantity']['high'] = entry['Q high']
        other_dict['quantity']['low'] = entry['Q low']
        if type(other_dict['quantity']['high']) == float:
            if math.isnan(other_dict['quantity']['high']):
                other_dict['quantity'] = {}
                other_dict['quantity']['baseline'] = entry['Quantity (Q)']
        other_dict['material_unit_price'] = {}
        other_dict['material_unit_price']['low'] = entry['Price low']
        other_dict['material_unit_price']['high'] = entry['Price high']
        other_dict['updatedOn'] = int(np.floor(time.time()))
        for mat in mat_lib:
            if mat['id'] == other_dict['material_id']:
                other_dict['material'] = mat
                break
        
        est_mat_lst.append(other_dict)
        
    return est_mat_lst

def make_est_spent_cat(excel_path, est_id, version):
    #spent_cat_df = pd.read_excel(excel_path, sheet_name='4 Spent Catalyst')
    metal_id_dict = get_ids('metal_id_dict')
    support_id_dict = get_ids('support_id_dict')
    hazard_id_dict = get_ids('hazard_id_dict')
    density_id_dict = get_ids('density_id_dict')
    spent_cat_dict = spent_cat_to_json(excel_path,None)
    spent_cat_dict = json.loads(spent_cat_dict)
    metal_dict = spent_cat_dict['spent_cat_metal']
    support_dict = spent_cat_dict['spent_cat_support']
    hazard_dict = spent_cat_dict['spent_cat_hazard']
    density_dict = spent_cat_dict['spent_cat_bulk_density']
    with open_workbook(excel_path) as wb:
        sheets = wb.sheets()
        for sheet in sheets:
            if sheet.name == '4 Spent Catalyst':
                spent_cat = sheet
                break
    spent_cat_metal_dict = {}
    spent_cat_metal_dict['id'] = gen_id()
    spent_cat_metal_dict['version'] = version
    spent_cat_metal_dict['updatedOn'] = int(np.floor(time.time()))
    spent_cat_metal_dict['estimate_id'] = est_id
    spent_cat_support_dict = {}
    spent_cat_support_dict['id'] = gen_id()
    spent_cat_support_dict['version'] = version
    spent_cat_support_dict['updatedOn'] = int(np.floor(time.time()))
    spent_cat_support_dict['estimate_id'] = est_id
    spent_cat_hazard_dict = {}
    spent_cat_hazard_dict['id'] = gen_id()
    spent_cat_hazard_dict['version'] = version
    spent_cat_hazard_dict['updatedOn'] = int(np.floor(time.time()))
    spent_cat_hazard_dict['estimate_id'] = est_id
    spent_cat_density_dict = {}
    spent_cat_density_dict['id'] = gen_id()
    spent_cat_density_dict['version'] = version
    spent_cat_density_dict['updatedOn'] = int(np.floor(time.time()))
    spent_cat_density_dict['estimate_id'] = est_id
    for rownum in range(spent_cat.nrows):
        row_value = spent_cat.row_values(rownum)
        if 'Metal to recover' in row_value:
            metal_name = locate_data(row_value, 'Metal to recover',False)
            if metal_name:
                spent_cat_metal_dict['spent_cat_metal_id'] = metal_id_dict[metal_name]
        elif 'Support' in row_value:
            support_name = locate_data(row_value, 'Support', False)
            if support_name:
                spent_cat_support_dict['spent_cat_support_id'] = support_id_dict[support_name]
        elif 'Metal wt. % of AP' in row_value:
            metal_weight = locate_data(row_value, 'Metal wt. % of AP')
            if any(metal_weight.values()):
                spent_cat_metal_dict['metal_weight'] = metal_weight
        elif 'Catalyst bulk density' in row_value:
            cat_bulk_density = locate_data(row_value, 'Catalyst bulk density')
            if any(cat_bulk_density.values()):
                spent_cat_metal_dict['cat_bulk_density'] = cat_bulk_density
        elif 'Has trace Sn, Cu, Fe > 2% of AP?' in row_value:
            has_trace_element = locate_data(row_value, 'Has trace Sn, Cu, Fe > 2% of AP?', False)
            if has_trace_element:
                spent_cat_metal_dict['has_trace_element'] = has_trace_element
        elif 'Metal content in fresh catalyst' in row_value:
            spent_cat_metal_dict['metal_content'] = locate_data(row_value, 'Metal content in fresh catalyst')
        elif 'Metal losses during refining (typical)' in row_value:
            spent_cat_metal_dict['loss'] = locate_data(row_value, 'Metal losses during refining (typical)')
        elif 'Spot price' in row_value:
            spent_cat_metal_dict['spot_price'] = locate_data(row_value, 'Spot price')
        elif 'Planned reactor configuration' in row_value:
            planned_use = locate_data(row_value, 'Planned reactor configuration',False)
            if planned_use:
                spent_cat_support_dict['planned_use'] = planned_use
        elif 'Catalyst solids after use' in row_value:
            spent_cat_support_dict['cat_solids_after_use'] = locate_data(row_value, 'Catalyst solids after use', False)
#        elif 'Metal losses during use (typical)' in row_value:
#            spent_cat_support_dict['metal_loss_during_use'] = locate_data(row_value, 'Metal losses during use (typical)', False)
        elif 'Classification for Sale or Landfill' in row_value:
            hazard_name = locate_data(row_value, 'Classification for Sale or Landfill',False)
            if hazard_name:
                spent_cat_hazard_dict['spent_cat_hazard_id'] = hazard_id_dict[hazard_name]
        #elif 'Landfill cost' in row_value:
        #    spent_cat_hazard_dict['landfill_fee'] = locate_data(row_value, 'Landfill cost')
        elif 'Catalyst/material type' in row_value:
            density_name = locate_data(row_value, 'Catalyst/material type', False)
            spent_cat_density_dict['spent_cat_bulk_density_id'] = density_id_dict[density_name]
            
    if metal_name:
        for metal in metal_dict:
            if metal['id'] == spent_cat_metal_dict['spent_cat_metal_id']:
                spent_cat_metal_dict['spent_cat_metal'] = metal
                break
        
    #spent_cat_metal_dict['price_per_mass_unit_catalyst'] = spent_cat.cell() not sure which value we want here
    #spent_cat_metal_dict['price_scaled'] again unsure of what value to pull here
    
    if support_name:
        for support in support_dict:
            if support['id'] == spent_cat_support_dict['spent_cat_support_id']:
                spent_cat_support_dict['spent_cat_support'] = support
                break
    
    if hazard_name:
        for hazard in hazard_dict:
            if hazard['id'] == spent_cat_hazard_dict['spent_cat_hazard_id']:
                spent_cat_hazard_dict['spent_cat_hazard'] = hazard
                break

    for density in density_dict:
        if density['id'] == spent_cat_density_dict['spent_cat_bulk_density_id']:
            spent_cat_density_dict['spent_cat_bulk_density'] = density
            break
#    return spent_cat_dict
    return spent_cat_metal_dict, spent_cat_support_dict, spent_cat_hazard_dict, spent_cat_density_dict
    
def make_est_process_utilities(excel_path, est_id, version, basis_unit):
    with open_workbook(excel_path) as wb:
        for sheet in wb.sheets():
            if sheet.name == '3c Utilities':
                utilities = sheet
                break
    for rownum in range(utilities.nrows):
        row_value = utilities.row_values(rownum)
        if 'Utility' in row_value:
            utility_start = rownum
            break
    utilities_df = pd.read_excel(excel_path, sheet_name='3c Utilities', 
                                    skiprows=utility_start, usecols='C:N')
    utilities_df = utilities_df[utilities_df['Unit Cost'] != "Totals"]
    utilities_df = utilities_df[utilities_df['Unit Cost'].notnull()]
    utility_lst = []
    for entry in utilities_df.iterrows():
        entry = entry[1]
        utility_dict = {}
        utility_dict['id'] = gen_id()
        utility_dict['estimate_id'] = est_id
        utility_dict['version'] = version
        utility_dict['name'] = entry['Utility']
        utility_dict['consumption'] = {}
        utility_dict['consumption']['baseline'] = entry['Consumption per %s catalyst' %basis_unit]
        utility_dict['consumption']['low'] = entry['Low']
        utility_dict['consumption']['high'] = entry['High']
        if type(utility_dict['consumption']['low']) == float:
            if math.isnan(utility_dict['consumption']['low']):
                utility_dict['consumption'] = {}
                utility_dict['consumption']['baseline'] = entry['Consumption per %s catalyst' %basis_unit]
        utility_dict['unit_cost'] = {}
        utility_dict['unit_cost']['baseline'] = entry['Unit Cost']
        utility_dict['unit_cost']['low'] = entry['Low.1']
        utility_dict['unit_cost']['high'] = entry['High.1']
        if type(utility_dict['unit_cost']['low']) == float:
            if math.isnan(utility_dict['unit_cost']['low']):
                utility_dict['unit_cost'] = {}
                utility_dict['unit_cost']['baseline'] = entry['Unit Cost']
        utility_lst.append(utility_dict)
    return utility_lst

def make_est_cap_ex(excel_path, est_id, version):
    with open_workbook(excel_path) as wb:
        for sheet in wb.sheets():
            if sheet.name == '3d CapEx':
                CapEx = sheet
                break
    for rownum in range(CapEx.nrows):
        row_value = CapEx.row_values(rownum)
        if 'Direct Capital' in row_value:
            direct_capital_start = rownum
        elif 'Indirect Capital' in row_value:
            indirect_capital_start = rownum
        elif 'Total Fixed Capital Investment (FCI)' in row_value:
            fci_start = rownum
        elif 'Total Direct' in row_value:
            direct_end = rownum - 1
        elif 'Total Indirect' in row_value:
            indirect_end = rownum - 1
        elif 'Total Capital Investment (TCI)' in row_value:
            fci_end = rownum
        
    direct_cap_ex_df = pd.read_excel(excel_path, sheet_name='3d CapEx', 
                                     skiprows=direct_capital_start, usecols='C:H', skipfooter=fci_end - direct_end)
    indirect_cap_ex_df = pd.read_excel(excel_path, sheet_name='3d CapEx', 
                                       skiprows=indirect_capital_start, usecols='C:H', skipfooter=fci_end - indirect_end)
    fci_df = pd.read_excel(excel_path, sheet_name='3d CapEx',
                           skiprows=fci_start, usecols='C:H', skipfooter=1)
    fci_df.columns = (['capital','baseline','low','high','percentpurchase','value'])
    factored_capital_costs = []
    for entry in direct_cap_ex_df.iterrows():
        entry = entry[1]
        cost_dict = {}
        cost_dict['id'] = gen_id()
        cost_dict['version'] = version
        cost_dict['updatedOn'] = int(np.floor(time.time()))
        cost_dict['name'] = entry['Direct Capital']
        cost_dict['percent_purchase_cost'] = {}
        cost_dict['percent_purchase_cost']['baseline'] = entry['Unnamed: 1']
        cost_dict['percent_purchase_cost']['low'] = entry['Unnamed: 2']
        cost_dict['percent_purchase_cost']['high'] = entry['Unnamed: 3']
        cost_dict['category'] = 'direct'
        cost_dict['estimate_id'] = est_id
        factored_capital_costs.append(cost_dict)
    for entry in indirect_cap_ex_df.iterrows():
        entry = entry[1]
        cost_dict = {}
        cost_dict['id'] = gen_id()
        cost_dict['version'] = version
        cost_dict['updatedOn'] = int(np.floor(time.time()))
        cost_dict['name'] = entry['Indirect Capital']
        cost_dict['percent_purchase_cost'] = {}
        cost_dict['percent_purchase_cost']['baseline'] = entry['Unnamed: 1']
        cost_dict['percent_purchase_cost']['low'] = entry['Unnamed: 2']
        cost_dict['percent_purchase_cost']['high'] = entry['Unnamed: 3']
        if cost_dict['name'] == "Working Capital":
            cost_dict['category'] = 'other'
        else:
            cost_dict['category'] = 'indirect'
        cost_dict['estimate_id'] = est_id
        factored_capital_costs.append(cost_dict)
    for entry in fci_df.iterrows():
        entry = entry[1]
        cost_dict = {}
        cost_dict['id'] = gen_id()
        cost_dict['version'] = version
        cost_dict['updatedOn'] = int(np.floor(time.time()))
        cost_dict['name'] = entry['capital']
        cost_dict['percent_purchase_cost'] = {}
        cost_dict['percent_purchase_cost']['baseline'] = entry['baseline']
        cost_dict['percent_purchase_cost']['low'] = entry['low']
        cost_dict['percent_purchase_cost']['high'] = entry['high']
        if cost_dict['name'] == "Working Capital":
            cost_dict['category'] = 'other'
        else:
            cost_dict['category'] = 'indirect'
        cost_dict['estimate_id'] = est_id
        factored_capital_costs.append(cost_dict)
    return factored_capital_costs

def make_est_op_ex(excel_path, est_id, version):
    with open_workbook(excel_path) as wb:
        for sheet in wb.sheets():
            if sheet.name == '3e OpEx':
                OpEx = sheet
                break
    for rownum in range(OpEx.nrows):
        row_value = OpEx.row_values(rownum)
        if 'Direct Operating Costs' in row_value:
            direct_op_start = rownum
        elif 'Fixed/Indirect Operating Costs' in row_value:
            fixed_op_start = rownum
        elif 'General Expenses' in row_value:
            gen_op_start = rownum
        elif 'Total: Labor, Supplies, Maintenance, Lab (LSM)' in row_value:
            direct_end = rownum - 1
        elif 'Total: Taxes, Insurance, Rent, Overhead' in row_value:
            fixed_end = rownum - 1
        elif 'Total: Admin, Dist., Mkting., R&D' in row_value:
            gen_end = rownum     
    factored_operating_costs = []
    direct_op_ex_df = pd.read_excel(excel_path, sheet_name='3e OpEx', 
                                    skiprows=direct_op_start, usecols='C:G', skipfooter=gen_end - direct_end)
    fixed_op_ex_df = pd.read_excel(excel_path, sheet_name='3e OpEx',
                                   skiprows=fixed_op_start, usecols='C:G', skipfooter=gen_end - fixed_end)
    gen_op_ex_df = pd.read_excel(excel_path, sheet_name='3e OpEx',
                                 skiprows=gen_op_start, usecols='C:G', skipfooter=1)
    for entry in direct_op_ex_df.iterrows():
        entry = entry[1]
        cost_dict = {}
        cost_dict['id'] = gen_id()
        cost_dict['version'] = version
        cost_dict['updatedOn'] = int(np.floor(time.time()))
        cost_dict['name'] = entry['Direct Operating Costs']
        cost_dict['factor'] = {}
        cost_dict['factor']['baseline'] = entry['Unnamed: 1']
        cost_dict['factor']['low'] = entry['Unnamed: 2']
        cost_dict['factor']['high'] = entry['Unnamed: 3']
        cost_dict['unit'] = entry['Unnamed: 4']
        if cost_dict['unit'] == '% of DL':
            cost_dict['_relative_path'] = 'estimate.direct_labor_total_cost_per_year'
        elif cost_dict['unit'] == '% of FCI':
            cost_dict['_relative_path'] = 'estimate.factored_capital_total_fixed_capital_investment'
        elif cost_dict['unit'] == '% of M&R':
            cost_dict['_relative_path'] = 'estimate.ref_operating_cost_maintenance_and_repair'
        cost_dict['category'] = 'direct'
        cost_dict['estimate_id'] = est_id
        factored_operating_costs.append(cost_dict)
    for entry in fixed_op_ex_df.iterrows():
        entry = entry[1]
        cost_dict = {}
        cost_dict['id'] = gen_id()
        cost_dict['version'] = version
        cost_dict['updatedOn'] = int(np.floor(time.time()))
        cost_dict['name'] = entry['Fixed/Indirect Operating Costs']
        cost_dict['factor'] = {}
        cost_dict['factor']['baseline'] = entry['Unnamed: 1']
        cost_dict['factor']['low'] = entry['Unnamed: 2']
        cost_dict['factor']['high'] = entry['Unnamed: 3']
        cost_dict['unit'] = entry['Unnamed: 4']
        if cost_dict['unit'] == '% of FCI':
            cost_dict['_relative_path'] = 'estimate.factored_capital_total_fixed_capital_investment'
        elif cost_dict['unit'] == '% of land':
            cost_dict['_relative_path'] = 'estimate.ref_capital_cost_land'
        elif cost_dict['unit'] == '% of LSM':
            cost_dict['_relative_path'] = 'estimate.direct_operating_total_cost'
        cost_dict['category'] = 'indirect'
        cost_dict['estimate_id'] = est_id
        factored_operating_costs.append(cost_dict)
    for entry in gen_op_ex_df.iterrows():
        entry = entry[1]
        cost_dict = {}
        cost_dict['id'] = gen_id()
        cost_dict['version'] = version
        cost_dict['updatedOn'] = int(np.floor(time.time()))
        cost_dict['name'] = entry['General Expenses']
        cost_dict['factor'] = {}
        cost_dict['factor']['baseline'] = entry['Unnamed: 1']
        cost_dict['factor']['low'] = entry['Unnamed: 2']
        cost_dict['factor']['high'] = entry['Unnamed: 3']
        cost_dict['unit'] = entry['Unnamed: 4']
        if cost_dict['unit'] == '% of op. costs':
            cost_dict['_relative_path'] = 'estimate.operating_costs'
        elif cost_dict['unit'] == '% of LSM':
            cost_dict['_relative_path'] = 'estimate.direct_operating_total_cost'
        cost_dict['category'] = 'other'
        if cost_dict['unit'] == "% of op. costs excluding PGM/noble metals content":
            cost_dict['_relative_path'] = 'estimate.material_total_cost_per_year'
            cost_dict['category'] = 'other'
        cost_dict['estimate_id'] = est_id
        factored_operating_costs.append(cost_dict)
    return factored_operating_costs

"""
The section below is used to add Ids to those objects which do not have ids specified
"""

def get_ids(lib):
    with open(os.getcwd() + '/all_ids.json') as f:
        json_ids_str = f.read()
        json_ids_dict = json.loads(json_ids_str)
    id_dict = json_ids_dict[lib]
    return id_dict

def add_id(lib, name):
    with open(os.getcwd() + '/all_ids.json','r') as f:
        json_ids_str = f.read()
        json_ids_dict = json.loads(json_ids_str)
        lib_to_edit = json_ids_dict[lib]
        new_id = gen_id()
        lib_to_edit[name] = new_id
        json_ids_dict[lib] = lib_to_edit
        json_ids_str = json.dumps(json_ids_dict, indent=2)
    with open(os.getcwd() + '/all_ids.json','w') as g:
        g.write(json_ids_str)
    return new_id
        
        